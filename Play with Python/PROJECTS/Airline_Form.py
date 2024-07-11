from email.headerregistry import Address
from http.client import OK
from lib2to3.pgen2.token import NAME
from sre_parse import State
from  tkinter import  ttk
import tkinter as tk 
from tkinter import Frame, Label
from tkinter import *
import openpyxl
from openpyxl import Workbook
import pathlib

from tkinter import messagebox
from django.forms import PasswordInput

#importing connection
#establishing connection
    

file=pathlib.Path("backenda.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]= "Name"
    sheet["B1"]="Contact"
    sheet["C1"]= "Dob"
    sheet["D1"]="City"
    sheet["E1"]= "State"
    sheet["F1"]="MotherNmae"
    sheet["G1"]= "FatherName"
    sheet["H1"]="Address"
    sheet["I1"]= "Password"
    sheet["J1"]="Email"
    sheet["K1"]= "RollNo"
    sheet["L1"]="Gender"
    
    file.save("backenda.xlsx")
def ok():
    a=e1.get()
    b=e2.get()
    c=e3.get()
    f=e6.get()
    d=city.get()
    e=state.get()
   
    #l=gender.get()
   
    
    print(a)
    print(b)
    print(c)
    print(d)
    print(e)
    print(f)
   
  #  print(l)
    file=openpyxl.load_workbook("backenda.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=a)
    sheet.cell(column=2,row=sheet.max_row,value=b)
    sheet.cell(column=3,row=sheet.max_row,value=c)
    sheet.cell(column=4,row=sheet.max_row,value=d)
    sheet.cell(column=5,row=sheet.max_row,value=e)
    sheet.cell(column=6,row=sheet.max_row,value=f)

    #sheet.cell(column=12,row=sheet.max_row,value=l)
    if gender.get()==1:
        gen="Male"
        print("Male")
        sheet.cell(column=12,row=sheet.max_row,value="Male")
    else:
            print("Female")
            sheet.cell(column=12,row=sheet.max_row,value="Female")
    file.save("backenda.xlsx")
root = Tk()
root.geometry("500x450")
root.title("Airlines Booking System")
R_label = Label(root, text="Airlines Booking System", width=20, font=("bold",20))
R_label.place(x=90, y=53)
FN_label = Label(root, text="Full Name", width=20, font=("bold",10))
FN_label.place(x=68, y=130)
e1= Entry(root)
e1.place(x=240, y=130)
E_label = Label(root, text="Enter your Aadhar no.", width=20, font=("bold",10))
E_label.place(x=68, y=160)
e2= Entry(root)
e2.place(x=240, y=160)
E_label = Label(root, text="Mobile no.", width=20, font=("bold",10))
E_label.place(x=68, y=190)
e3= Entry(root)
e3.place(x=240, y=190)
E_label = Label(root, text="Enter your age", width=20, font=("bold",10))
E_label.place(x=68, y=220)
e6= Entry(root)
e6.place(x=240, y=220)
G_label = Label(root, text="Gender", width=20, font=("bold",10))
G_label.place(x=70, y=250)
gender = IntVar()
Radiobutton(root, text="Male", padx=5, variable=gender, value=1).place(x=235, y=250)
Radiobutton(root, text="Female", padx=5, variable=gender, value=2).place(x=290, y=250)
C_label = Label(root, text="Select your Boarding", width=20, font=("bold",10))
C_label.place(x=70, y=280)
list1 = ['Mumbai', 'Delhi', 'Nanded', 'Latur', 'Pune' ]
city = StringVar()
droplist = OptionMenu(root, city, *list1)
droplist.config(width=15)
city.set("Boarding")
droplist.place(x=240, y=280)
C_label = Label(root, text="Select your destination", width=20, font=("bold",10))
C_label.place(x=70, y=310)
list1 = ['Canada', 'Japan', 'UK', 'South Africa', 'Nepal' ]
state = StringVar()
droplist = OptionMenu(root, state, *list1)
droplist.config(width=15)
state.set("Destination")
droplist.place(x=240, y=310)
P_label = Label(root, text="Check it", width=20, font=("bold",10))
P_label.place(x=85, y=340)
var1 = IntVar()
Checkbutton(root, text="I accept all the rules of Airlines", padx=5, variable=var1).place(x=235, y=340)
Button(root, text="Submit",command=ok, width=20, bg='brown', fg='white').place(x=180, y=380)
root.mainloop()

